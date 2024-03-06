#C:\\Users\\m\\Desktop\\商贸学院软件设计大赛\\比赛数据\\部分识别样本\\昆虫编号.xlsx


# 导入pandas库，简写为pd
import openpyxl as vb
 
# 创建一个工作表和工作簿
wb = vb.load_workbook('C:\\Users\\m\\Desktop\\商贸学院软件设计大赛\\比赛数据\\部分识别样本\\昆虫编号.xlsx')
sheet = wb["Sheet1"]    # 获取的表名为Sheet
 
# 按列获取单元格,min_col：读取表格的1，2列，min_row：从第二行开始读到第20行，所有内容
# 读取顺序是一列再一列
#读取特定列
# cols = sheet.iter_cols(min_col=7, max_col=7, min_row=2, max_row=100)
# # 打印读取内容
# table=[]
# for col in cols:
#     for cell in col:
#         #print(cell.value)
#         table.append(cell.value)


# print(table[10])

 
# 读取指定列,先用列表推导式生成包含每一列中所有单元格的元组的列表，再对列表取索引
# one_cols = [val for val in sheet.columns][1]    # 获取第二列
# # 打印读取内容
# for col in one_cols:
#     print(col.value)


#读取特定行
clo=cols = sheet.iter_cols(min_col=0, max_col=10, min_row=2, max_row=2)
for col in cols:
    for cell in col:
        print(cell.value)
        